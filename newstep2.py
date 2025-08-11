print('Hello! Welcome!')

import os
import re
import pandas as pd
import numpy as np
from datetime import datetime
import holidays
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

directory = os.path.dirname(__file__)
print(directory)

# 查找error文件
error_file = None
time_range = None
for file in os.listdir(directory):
    print(file)
    if 'error' in file:
        import_path = os.path.join(directory, file)
        print(import_path)
        time_range = file.split("(")[1].split(")")[0]
        df = pd.read_excel(import_path)
        df_new = df.copy()
        error_file = file
        break

if error_file is None:
    print("未找到error文件!")
    quit()

file_name = 'work_attendance' + '(' + time_range + ')' + '.xlsx'
final_output_path = os.path.join(directory, file_name)
print(final_output_path)

# 删除已存在的输出文件
if os.path.exists(final_output_path):
    os.remove(final_output_path)


def daily_working_time(i):
    time_part = []

    while (i >= 1):
        time_difference = (time_list_normalized[i - 1] - time_list_normalized[i - 2]).total_seconds() / (60 * 60)
        i = i - 2
        time_part.append(time_difference)
    daily_time = round(sum(time_part), 2)

    return daily_time


# 在任何处理之前先保存原始打卡时间数据
df_original_times = df.copy()

# 将整个数据框转换为字符串
df = df.astype(str)

# 记录有问题的数据
problematic_data = []
problematic_cells = []  # 记录有问题的单元格位置 (行, 列)

# 将时间快照转换为每日总工作时间
for i in range(df.shape[0]):
    for j in range(len(df.columns) - 1):
        if (df.iloc[i, j + 1] == 'nan'):
            continue
        else:
            # 分割字符串 - 处理多种分隔符
            raw_time_str = str(df.iloc[i, j + 1])

            # 尝试多种分割方式
            time_list = []
            if '\n' in raw_time_str:
                time_list = raw_time_str.split('\n')
            else:
                # 使用正则表达式提取所有时间格式 HH:MM
                import re

                time_pattern = r'\d{1,2}:\d{2}'
                time_list = re.findall(time_pattern, raw_time_str)
                print(f"正则提取时间: {raw_time_str} -> {time_list}")  # 调试信息

            # 清理时间列表
            time_list = [t.strip() for t in time_list if t.strip() and t.strip() != '']

            # 将字符串规范化为日期
            time_list_normalized = []
            valid_times = True
            for time_str in time_list:
                try:
                    # 确保时间格式正确
                    if ':' in time_str and len(time_str.split(':')) == 2:
                        hour, minute = time_str.split(':')
                        # 验证小时和分钟的有效性
                        if 0 <= int(hour) <= 23 and 0 <= int(minute) <= 59:
                            date_time_obj = datetime.strptime(time_str, '%H:%M')
                            time_list_normalized.append(date_time_obj)
                        else:
                            print(f"⚠️ 无效时间格式: {time_str} (员工: {df.iloc[i, 0]}, 列: {j + 1})")
                            problematic_data.append(f"无效时间 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间: {time_str}")
                            problematic_cells.append((i, j + 1))  # 记录单元格位置
                            valid_times = False
                    else:
                        print(f"⚠️ 跳过无效时间格式: {time_str} (员工: {df.iloc[i, 0]}, 列: {j + 1})")
                        problematic_data.append(f"格式错误 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间: {time_str}")
                        problematic_cells.append((i, j + 1))  # 记录单元格位置
                        valid_times = False
                except ValueError as e:
                    print(f"⚠️ 时间解析错误 '{time_str}': {e} (员工: {df.iloc[i, 0]}, 列: {j + 1})")
                    problematic_data.append(
                        f"解析错误 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间: {time_str}, 错误: {e}")
                    problematic_cells.append((i, j + 1))  # 记录单元格位置
                    valid_times = False
                    continue

            # 只有当所有时间都有效时才进行计算
            if valid_times and len(time_list_normalized) > 0:
                # 双重检查单日时间快照的奇偶性并计算每日总工作时间
                if (len(time_list_normalized) % 2 == 0):
                    df_new.iloc[i, j + 1] = daily_working_time(len(time_list_normalized))
                else:
                    print(
                        f"⚠️ 奇数时间记录 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间数量: {len(time_list_normalized)}")
                    problematic_data.append(f"奇数时间记录 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间: {time_list}")
                    problematic_cells.append((i, j + 1))  # 记录单元格位置
                    df_new.iloc[i, j + 1] = 0  # 设置为0而不是停止程序
            else:
                # 有问题的数据设置为0，并记录位置
                if not valid_times:
                    # problematic_cells.append((i, j+1))  # 已经在上面记录了
                    pass
                df_new.iloc[i, j + 1] = 0

# 计算第一周的工作时间
total1 = df_new.iloc[:, 1:8].sum(axis=1).to_list()
n = len(total1)
HEG1 = [0] * n

for i in range(n):
    if total1[i] > 0:
        if total1[i] > 40:
            HEG1[i] = 40
        else:
            HEG1[i] = total1[i]
    else:
        continue

OT1 = [0] * n

for i in range(n):
    if total1[i] > 0:
        if total1[i] > 40:
            OT1[i] = total1[i] - 40
        else:
            OT1[i] = 0
    else:
        continue

# 计算第二周的工作时间
total2 = df_new.iloc[:, 8:17].sum(axis=1).to_list()
HEG2 = [0] * n

for i in range(n):
    if total2[i] > 0:
        if total2[i] > 40:
            HEG2[i] = 40
        else:
            HEG2[i] = total2[i]
    else:
        continue

OT2 = [0] * n

for i in range(n):
    if total2[i] > 0:
        if total2[i] > 40:
            OT2[i] = total2[i] - 40
        else:
            OT2[i] = 0
    else:
        continue

Total_HEG = [HEG1[i] + HEG2[i] for i in range(n)]
Total_OT = [OT1[i] + OT2[i] for i in range(n)]

df_new.insert(8, "HEG1", HEG1)
df_new.insert(9, "OT1", OT1)
df_new.insert(17, "HEG2", HEG2)
df_new.insert(18, "OT2", OT2)
df_new.insert(19, "Total_HEG", Total_HEG)
df_new.insert(20, "Total_OT", Total_OT)
metrics_list = [8, 9, 17, 18, 19, 20]

name_list = df_new[(df_new['HEG1'] > 30) | (df_new['HEG2'] > 30)]['name'].to_list()

df_new = df_new.replace(0, '').infer_objects(copy=False)

# 处理原始时间数据用于显示
df_original_for_display = df_original_times.astype(str).replace('nan', '')

r = df_original_for_display.shape[0]
c = df_original_for_display.shape[1]

highlight_cols_m = []  # 迟到
highlight_cols_n = []  # 中午不打卡
highlight_cols_e = []  # 早退

for i in range(c - 1):
    highlight_rows_m = []
    highlight_rows_n = []
    highlight_rows_e = []
    for j in range(r):
        if df_original_for_display.iloc[j, 0] in name_list:
            if (df_original_for_display.iloc[j, i + 1] == ''):
                continue
            else:
                # 使用与前面相同的时间解析逻辑
                raw_time_str = str(df_original_for_display.iloc[j, i + 1])

                # 提取时间字符串
                time_snapshot = []
                if '\n' in raw_time_str:
                    time_snapshot = raw_time_str.split('\n')
                else:
                    import re

                    time_pattern = r'\d{1,2}:\d{2}'
                    time_snapshot = re.findall(time_pattern, raw_time_str)

                # 清理时间列表，验证时间格式
                valid_times = []
                for t in time_snapshot:
                    t = t.strip()
                    if t and ':' in t:
                        try:
                            hour, minute = t.split(':')
                            if 0 <= int(hour) <= 23 and 0 <= int(minute) <= 59:
                                valid_times.append(t)
                        except:
                            continue

                time_snapshot = valid_times

                if len(time_snapshot) == 0:
                    continue

                try:
                    check_in_time = datetime.strptime(time_snapshot[0], '%H:%M')
                    morning_reference_time = datetime.strptime('10:00', '%H:%M')
                    check_out_time = datetime.strptime(time_snapshot[-1], '%H:%M')
                    evening_reference_time = datetime.strptime('17:00', '%H:%M')
                    check_times = len(time_snapshot)
                except ValueError as e:
                    print(
                        f"⚠️ 迟到早退检测 - 时间解析错误: 员工: {df_original_for_display.iloc[j, 0]}, 时间: {time_snapshot}, 错误: {e}")
                    problematic_data.append(
                        f"迟到早退检测错误 - 员工: {df_original_for_display.iloc[j, 0]}, 列: {i + 1}, 时间: {time_snapshot}, 错误: {e}")
                    continue

                if (check_in_time > morning_reference_time):
                    highlight_rows_m.append(j)
                    if (check_times == 2):
                        highlight_rows_n.append(j)
                        if (check_out_time < evening_reference_time):
                            highlight_rows_e.append(j)
                        else:
                            continue
                    elif (check_out_time < evening_reference_time):
                        highlight_rows_e.append(j)
                    else:
                        continue
                elif (check_times == 2):
                    highlight_rows_n.append(j)
                    if (check_out_time < evening_reference_time):
                        highlight_rows_e.append(j)
                    else:
                        continue
                elif (check_out_time < evening_reference_time):
                    highlight_rows_e.append(j)
                else:
                    continue
        else:
            continue
    highlight_cols_m.append(highlight_rows_m)
    highlight_cols_n.append(highlight_rows_n)
    highlight_cols_e.append(highlight_rows_e)

# 创建最终显示的数据框，先从原始时间开始
df_final = df_original_for_display.copy()

# 获取原始日期列的数量
original_date_cols = len(df_final.columns) - 1  # 减去name列

# 在原始日期列之后添加计算的工作小时数列
for i in range(original_date_cols):
    col_name = f"{df_final.columns[i + 1]}_小时"
    df_final.insert(i + 1 + original_date_cols, col_name, df_new.iloc[:, i + 1])

# 然后在最后添加统计列
df_final["HEG1"] = HEG1
df_final["OT1"] = OT1
df_final["HEG2"] = HEG2
df_final["OT2"] = OT2
df_final["Total_HEG"] = Total_HEG
df_final["Total_OT"] = Total_OT

# 更新metrics_list以反映新的列位置 - 指向统计列
metrics_start = len(df_final.columns) - 6  # 最后6列是统计列
metrics_list = list(range(metrics_start, len(df_final.columns)))

# 将0替换为空字符串（仅在小时数和统计列中），使用infer_objects()避免警告
hours_and_stats_cols = list(range(1 + original_date_cols, len(df_final.columns)))
for col_idx in hours_and_stats_cols:
    df_final.iloc[:, col_idx] = df_final.iloc[:, col_idx].replace(0, '').infer_objects(copy=False)

# 获取美国假期
US_holidays = pd.DataFrame.from_dict(holidays.US(years=2022).items())
US_holidays.columns = ["date", "holiday_name"]

# 定义公司假期
my_vacation = ["New Year's Day", "Independence Day", "Labor Day", "Thanksgiving", "Christmas Day"]

temp_list = []
for i in US_holidays['holiday_name']:
    if i in my_vacation:
        x = US_holidays[US_holidays['holiday_name'] == i].index[0]
        temp_list.append(x)

my_vacation_date = list(US_holidays.iloc[temp_list, 0])

time_str = time_range.split("-")

# 检查时间范围是否跨越两年
if len(time_str[0]) != len(time_str[1]):
    year1 = time_str[0][:4]
    start_day = time_str[0][-4:]
    end_day = time_str[1][-4:]
    start_date = datetime.strptime(year1 + start_day, '%Y%m%d').date()
    end_date = datetime.strptime(year1 + end_day, '%Y%m%d').date()
else:
    year1 = time_str[0][:4]
    year2 = time_str[1][:4]
    start_day = time_str[0][-4:]
    end_day = time_str[1][-4:]
    start_date = datetime.strptime(year1 + start_day, '%Y%m%d').date()
    end_date = datetime.strptime(year2 + end_day, '%Y%m%d').date()

# 检查时间范围内是否有假期
holiday_column = None
for time in my_vacation_date:
    if start_date < time < end_date:
        index = US_holidays[US_holidays['date'] == time].index[0]
        holiday = US_holidays.holiday_name[index]
        # 更改假期列的名称（在原始时间列中）
        if time.day in df_final.columns:
            df_final = df_final.rename(columns={time.day: holiday})
            # 找到要高亮显示的列名索引
            holiday_column = df_final.columns.get_loc(holiday)
    else:
        continue

# 使用openpyxl创建多工作表Excel文件
print("开始创建Excel文件...")
print(f"df_final 列名: {list(df_final.columns)}")
print(f"df_final 形状: {df_final.shape}")
print(f"原始时间列数: {original_date_cols}")
print(f"统计列位置: {metrics_list}")

from openpyxl import Workbook

workbook = Workbook()
workbook.remove(workbook.active)  # 移除默认工作表

# 创建工作表 - 第一个工作表使用包含原始时间+计算小时数的数据框，其他工作表使用原始时间
sheet_names = ["时间汇总", "迟到", "中午不打卡", "早退"]
sheets_data = [df_final, df_original_for_display, df_original_for_display, df_original_for_display]

for sheet_name, data in zip(sheet_names, sheets_data):
    ws = workbook.create_sheet(title=sheet_name)

    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

# 定义颜色
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
problem_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 深红色用于问题数据

print("开始高亮显示 迟到|中午不打卡|早退...")
print("请稍候...")

# 高亮时间汇总工作表
sheet1 = workbook["时间汇总"]
# 调整metrics_list，因为Excel列索引从1开始
for i in metrics_list:
    sheet1.cell(row=1, column=i + 1).fill = yellow_fill

sheet1.cell(row=1, column=1).fill = red_fill
if holiday_column is not None:
    sheet1.cell(row=1, column=holiday_column + 1).fill = green_fill

# 标红有问题的原始时间数据单元格（在原始时间列中）
for row_idx, col_idx in problematic_cells:
    if col_idx <= original_date_cols:  # 只在原始时间列中标红
        sheet1.cell(row=row_idx + 2, column=col_idx + 1).fill = problem_fill

# 高亮迟到工作表
sheet2 = workbook["迟到"]
sheet2.cell(row=1, column=1).fill = red_fill
for i in range(c - 1):
    for j in highlight_cols_m[i]:
        sheet2.cell(row=j + 2, column=i + 2).fill = red_fill

# 标红有问题的数据单元格
for row_idx, col_idx in problematic_cells:
    if col_idx <= c - 1:
        sheet2.cell(row=row_idx + 2, column=col_idx + 1).fill = problem_fill

# 高亮中午不打卡工作表
sheet3 = workbook["中午不打卡"]
sheet3.cell(row=1, column=1).fill = red_fill
for i in range(c - 1):
    for j in highlight_cols_n[i]:
        sheet3.cell(row=j + 2, column=i + 2).fill = red_fill

# 标红有问题的数据单元格
for row_idx, col_idx in problematic_cells:
    if col_idx <= c - 1:
        sheet3.cell(row=row_idx + 2, column=col_idx + 1).fill = problem_fill

# 高亮早退工作表
sheet4 = workbook["早退"]
sheet4.cell(row=1, column=1).fill = red_fill
for i in range(c - 1):
    for j in highlight_cols_e[i]:
        sheet4.cell(row=j + 2, column=i + 2).fill = red_fill

# 标红有问题的数据单元格
for row_idx, col_idx in problematic_cells:
    if col_idx <= c - 1:
        sheet4.cell(row=row_idx + 2, column=col_idx + 1).fill = problem_fill

# 自动调整列宽
for sheet in workbook.worksheets:
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

workbook.save(final_output_path)
workbook.close()

print(f"处理完成! 输出文件: {final_output_path}")
print(f"生成了 {len(sheet_names)} 个工作表:")
for sheet_name in sheet_names:
    print(f"  - {sheet_name}")
print("高亮显示说明:")
print("  - 🟡 黄色背景: 工时统计列")
print("  - 🔴 浅红色背景: 迟到/中午不打卡/早退")
print("  - 🟢 绿色背景: 假期列")
print("  - 🔥 深红色背景: 有问题的时间数据 (需要人工确认)")
print(f"  - 共发现 {len(problematic_cells)} 个问题单元格已标红")

# 显示有问题的数据总结
if problematic_data:
    print("\n" + "=" * 50)
    print("⚠️  发现以下有问题的数据:")
    print("=" * 50)
    for i, problem in enumerate(problematic_data, 1):
        print(f"{i}. {problem}")
    print("=" * 50)
    print(f"总共发现 {len(problematic_data)} 个问题")
    print("这些问题数据已被设置为0，程序继续执行")
    print("🔥 在Excel文件中，有问题的单元格已用深红色标记，请人工确认并修正")
    print("💡 双击有问题的单元格可以查看并编辑原始数据")
else:
    print("\n✅ 没有发现问题数据，所有时间记录都正常处理")