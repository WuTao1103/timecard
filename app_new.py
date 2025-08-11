from flask import Flask, request, jsonify, send_file, render_template_string
import os
import pandas as pd
import numpy as np
from datetime import datetime
import holidays
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import uuid
import traceback
import re
import json

app = Flask(__name__)

# 添加CORS支持
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# 配置
UPLOAD_FOLDER = './uploads'
PROCESSED_FOLDER = './processed'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

def get_minimum_distance(letter):
    """计算冒号之间的最小距离"""
    position = []
    for i in range(len(letter)):
        if letter[i] == ':':
            position.append(i)

    if len(position) > 1:
        j = len(position) - 1
        distance = []
        while j >= 1:
            d = position[j] - position[j - 1]
            j = j - 1
            distance.append(d)
        return min(distance)
    else:
        return None

def daily_working_time(time_list_normalized):
    """计算每日工作时间"""
    time_part = []
    i = len(time_list_normalized)

    while (i >= 1):
        time_difference = (time_list_normalized[i - 1] - time_list_normalized[i - 2]).total_seconds() / (60 * 60)
        i = i - 2
        time_part.append(time_difference)

    return round(sum(time_part), 2)

def process_timecard_step1(file_path):
    """Step1处理逻辑 - 包含完整的错误检测和高亮功能"""
    try:
        df = pd.read_excel(file_path)

        # 获取时间范围
        time_range = (df.iloc[1, 2]).replace("/", "").replace("~", "-").replace(" ", "")

        total_rows = df.shape[0]
        employee_amount = int((total_rows - 2) / 3)

        date_row = df.iloc[2].to_list()
        date_range = [x for x in date_row if str(x) != 'nan']
        columns_name = list(map(int, date_range))
        columns_name.insert(0, 'name')

        df_new = pd.DataFrame(index=range(employee_amount), columns=columns_name)

        # 创建新的员工和日常检查表
        for i in range(employee_amount):
            df_new.iloc[i, 0] = df.iloc[(i + 1) * 3, 10]
            df_new.iloc[i, 1:] = df.iloc[(3 * (i + 1) + 1), 0:len(date_range)]

        df_new['nan_count'] = df_new.isna().sum(axis=1)
        df_new_sorted = df_new.sort_values(by='nan_count', ascending=True).reset_index(drop=True)
        df_new = df_new_sorted.drop('nan_count', axis=1)

        # 检测错误值位置
        error_value_location = []
        for i in range(employee_amount):
            for j in range(len(date_range)):
                string = str(df_new.iloc[i, j + 1])

                if string == 'nan':
                    continue
                else:
                    letter = [x for x in string]
                    min_distance = get_minimum_distance(letter)
                    if min_distance == 3:
                        location = [i, j + 1]
                        error_value_location.append(location)

        # 检测奇数时间记录
        highlight_index = []
        for i in range(len(date_range)):
            time_rows = df_new[((df_new.iloc[:, i + 1].str.count(":") % 2) == 1).values].index.to_list()
            highlight_index.append(time_rows)

        # 合并错误位置
        for i in range(len(error_value_location)):
            r = error_value_location[i][0]
            c = error_value_location[i][1]
            highlight_index[c - 1].append(r)

        # 保存文件并添加高亮
        output_filename = f'table_with_error_cells({time_range}).xlsx'
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)
        df_new.to_excel(output_path, index=None, header=True)

        # 使用openpyxl添加高亮显示
        workbook = load_workbook(output_path)
        worksheet = workbook.active
        highlight_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # 高亮有问题的单元格
        for i in range(len(date_range)):
            for j in highlight_index[i]:
                cell = worksheet.cell(row=j + 2, column=i + 2)
                cell.fill = highlight_fill

        workbook.save(output_path)
        workbook.close()

        # 生成错误报告
        error_details = []
        total_highlighted = sum(len(rows) for rows in highlight_index)

        if error_value_location:
            error_details.append(f"发现 {len(error_value_location)} 个冒号距离异常的时间记录")

        if total_highlighted > len(error_value_location):
            odd_time_count = total_highlighted - len(error_value_location)
            error_details.append(f"发现 {odd_time_count} 个奇数时间记录")

        return {
            'success': True,
            'time_range': time_range,
            'output_file': output_filename,
            'employee_count': employee_amount,
            'error_count': len(error_value_location),
            'total_highlighted': total_highlighted,
            'error_details': error_details
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }

def process_timecard_step2(error_file_path, time_range):
    """Step2处理逻辑 - 包含完整的工时计算、错误处理和多工作表生成"""
    try:
        df = pd.read_excel(error_file_path)
        df_new = df.copy()

        # 保存原始打卡时间数据
        df_original_times = df.copy()

        # 转换为字符串
        df = df.astype(str)

        # 记录有问题的数据
        problematic_data = []
        problematic_cells = []

        # 处理时间数据
        for i in range(df.shape[0]):
            for j in range(len(df.columns) - 1):
                if (df.iloc[i, j + 1] == 'nan'):
                    continue

                raw_time_str = str(df.iloc[i, j + 1])

                # 分割时间字符串
                time_list = []
                if '\n' in raw_time_str:
                    time_list = raw_time_str.split('\n')
                else:
                    time_pattern = r'\d{1,2}:\d{2}'
                    time_list = re.findall(time_pattern, raw_time_str)

                time_list = [t.strip() for t in time_list if t.strip()]

                # 验证和规范化时间
                time_list_normalized = []
                valid_times = True

                for time_str in time_list:
                    try:
                        if ':' in time_str and len(time_str.split(':')) == 2:
                            hour, minute = time_str.split(':')
                            if 0 <= int(hour) <= 23 and 0 <= int(minute) <= 59:
                                date_time_obj = datetime.strptime(time_str, '%H:%M')
                                time_list_normalized.append(date_time_obj)
                            else:
                                problematic_data.append(
                                    f"无效时间 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间: {time_str}")
                                problematic_cells.append((i, j + 1))
                                valid_times = False
                        else:
                            problematic_data.append(f"格式错误 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间: {time_str}")
                            problematic_cells.append((i, j + 1))
                            valid_times = False
                    except ValueError as e:
                        problematic_data.append(f"解析错误 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间: {time_str}")
                        problematic_cells.append((i, j + 1))
                        valid_times = False

                # 计算工作时间
                if valid_times and len(time_list_normalized) > 0:
                    if len(time_list_normalized) % 2 == 0:
                        df_new.iloc[i, j + 1] = daily_working_time(time_list_normalized)
                    else:
                        problematic_data.append(
                            f"奇数时间记录 - 员工: {df.iloc[i, 0]}, 列: {j + 1}, 时间数量: {len(time_list_normalized)}")
                        problematic_cells.append((i, j + 1))
                        df_new.iloc[i, j + 1] = 0
                else:
                    df_new.iloc[i, j + 1] = 0

        # 计算工时统计
        n = len(df_new)

        # 第一周工时计算
        total1 = df_new.iloc[:, 1:8].sum(axis=1).to_list()
        HEG1 = [min(40, max(0, t)) if t > 0 else 0 for t in total1]
        OT1 = [max(0, t - 40) if t > 40 else 0 for t in total1]

        # 第二周工时计算
        total2 = df_new.iloc[:, 8:17].sum(axis=1).to_list()
        HEG2 = [min(40, max(0, t)) if t > 0 else 0 for t in total2]
        OT2 = [max(0, t - 40) if t > 40 else 0 for t in total2]

        # 总计
        Total_HEG = [HEG1[i] + HEG2[i] for i in range(n)]
        Total_OT = [OT1[i] + OT2[i] for i in range(n)]

        # 插入统计列
        df_new.insert(8, "HEG1", HEG1)
        df_new.insert(9, "OT1", OT1)
        df_new.insert(17, "HEG2", HEG2)
        df_new.insert(18, "OT2", OT2)
        df_new.insert(19, "Total_HEG", Total_HEG)
        df_new.insert(20, "Total_OT", Total_OT)

        # 识别需要检查迟到早退的员工
        name_list = df_new[(df_new['HEG1'] > 30) | (df_new['HEG2'] > 30)]['name'].to_list()

        # 处理原始时间数据用于显示
        df_original_for_display = df_original_times.astype(str).replace('nan', '')

        # 创建最终显示的数据框
        df_final = df_original_for_display.copy()
        original_date_cols = len(df_final.columns) - 1

        # 添加计算的工作小时数列
        for i in range(original_date_cols):
            col_name = f"{df_final.columns[i + 1]}_小时"
            df_final.insert(i + 1 + original_date_cols, col_name, df_new.iloc[:, i + 1])

        # 添加统计列
        df_final["HEG1"] = HEG1
        df_final["OT1"] = OT1
        df_final["HEG2"] = HEG2
        df_final["OT2"] = OT2
        df_final["Total_HEG"] = Total_HEG
        df_final["Total_OT"] = Total_OT

        # 替换0为空字符串
        hours_and_stats_cols = list(range(1 + original_date_cols, len(df_final.columns)))
        for col_idx in hours_and_stats_cols:
            df_final.iloc[:, col_idx] = df_final.iloc[:, col_idx].replace(0, '')

        # 检测迟到早退
        r = df_original_for_display.shape[0]
        c = df_original_for_display.shape[1]

        highlight_cols_m = []  # 迟到
        highlight_cols_n = []  # 中午不打卡
        highlight_cols_e = []  # 早退

        attendance_issues = []

        for i in range(c - 1):
            highlight_rows_m = []
            highlight_rows_n = []
            highlight_rows_e = []

            for j in range(r):
                if df_original_for_display.iloc[j, 0] in name_list:
                    if df_original_for_display.iloc[j, i + 1] == '':
                        continue

                    raw_time_str = str(df_original_for_display.iloc[j, i + 1])

                    # 提取时间字符串
                    time_snapshot = []
                    if '\n' in raw_time_str:
                        time_snapshot = raw_time_str.split('\n')
                    else:
                        time_pattern = r'\d{1,2}:\d{2}'
                        time_snapshot = re.findall(time_pattern, raw_time_str)

                    # 验证时间格式
                    valid_times = []
                    for t in time_snapshot:
                        t = t.strip()
                        if t and ':' in t:
                            try:
                                hour, minute = t.split(':')
                                if 0 <= int(hour) <= 23 and 0 <= int(minute) <= 59:
                                    valid_times.append(t)
                            except:
                                continue

                    time_snapshot = valid_times

                    if len(time_snapshot) == 0:
                        continue

                    try:
                        check_in_time = datetime.strptime(time_snapshot[0], '%H:%M')
                        morning_reference_time = datetime.strptime('10:00', '%H:%M')
                        check_out_time = datetime.strptime(time_snapshot[-1], '%H:%M')
                        evening_reference_time = datetime.strptime('17:00', '%H:%M')
                        check_times = len(time_snapshot)

                        employee_name = df_original_for_display.iloc[j, 0]
                        date_col = df_final.columns[i + 1]

                        # 检测迟到
                        if check_in_time > morning_reference_time:
                            highlight_rows_m.append(j)
                            attendance_issues.append(
                                f"迟到 - {employee_name}, {date_col}, 上班时间: {time_snapshot[0]}")

                        # 检测中午不打卡
                        if check_times == 2:
                            highlight_rows_n.append(j)
                            attendance_issues.append(
                                f"中午不打卡 - {employee_name}, {date_col}, 打卡次数: {check_times}")

                        # 检测早退
                        if check_out_time < evening_reference_time:
                            highlight_rows_e.append(j)
                            attendance_issues.append(
                                f"早退 - {employee_name}, {date_col}, 下班时间: {time_snapshot[-1]}")

                    except ValueError as e:
                        problematic_data.append(
                            f"迟到早退检测错误 - 员工: {df_original_for_display.iloc[j, 0]}, 列: {i + 1}, 时间: {time_snapshot}")

            highlight_cols_m.append(highlight_rows_m)
            highlight_cols_n.append(highlight_rows_n)
            highlight_cols_e.append(highlight_rows_e)

        # 处理假期
        US_holidays = pd.DataFrame.from_dict(holidays.US(years=2022).items())
        US_holidays.columns = ["date", "holiday_name"]
        my_vacation = ["New Year's Day", "Independence Day", "Labor Day", "Thanksgiving", "Christmas Day"]

        temp_list = []
        for i in US_holidays['holiday_name']:
            if i in my_vacation:
                x = US_holidays[US_holidays['holiday_name'] == i].index[0]
                temp_list.append(x)

        my_vacation_date = list(US_holidays.iloc[temp_list, 0])

        # 解析时间范围
        time_str = time_range.split("-")
        if len(time_str[0]) != len(time_str[1]):
            year1 = time_str[0][:4]
            start_day = time_str[0][-4:]
            end_day = time_str[1][-4:]
            start_date = datetime.strptime(year1 + start_day, '%Y%m%d').date()
            end_date = datetime.strptime(year1 + end_day, '%Y%m%d').date()
        else:
            year1 = time_str[0][:4]
            year2 = time_str[1][:4]
            start_day = time_str[0][-4:]
            end_day = time_str[1][-4:]
            start_date = datetime.strptime(year1 + start_day, '%Y%m%d').date()
            end_date = datetime.strptime(year2 + end_day, '%Y%m%d').date()

        # 检查假期
        holiday_column = None
        for time in my_vacation_date:
            if start_date < time < end_date:
                index = US_holidays[US_holidays['date'] == time].index[0]
                holiday = US_holidays.holiday_name[index]
                if time.day in df_final.columns:
                    df_final = df_final.rename(columns={time.day: holiday})
                    holiday_column = df_final.columns.get_loc(holiday)

        # 创建Excel文件
        output_filename = f'work_attendance({time_range}).xlsx'
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)

        workbook = Workbook()
        workbook.remove(workbook.active)

        # 创建工作表
        sheet_names = ["时间汇总", "迟到", "中午不打卡", "早退"]
        sheets_data = [df_final, df_original_for_display, df_original_for_display, df_original_for_display]

        for sheet_name, data in zip(sheet_names, sheets_data):
            ws = workbook.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # 定义颜色
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        problem_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # 高亮时间汇总工作表
        sheet1 = workbook["时间汇总"]
        metrics_start = len(df_final.columns) - 6
        for i in range(metrics_start, len(df_final.columns)):
            sheet1.cell(row=1, column=i + 1).fill = yellow_fill

        sheet1.cell(row=1, column=1).fill = red_fill
        if holiday_column is not None:
            sheet1.cell(row=1, column=holiday_column + 1).fill = green_fill

        # 标红有问题的数据
        for row_idx, col_idx in problematic_cells:
            if col_idx <= original_date_cols:
                sheet1.cell(row=row_idx + 2, column=col_idx + 1).fill = problem_fill

        # 高亮其他工作表
        for sheet_idx, (sheet_name, highlight_cols) in enumerate(
                zip(["迟到", "中午不打卡", "早退"], [highlight_cols_m, highlight_cols_n, highlight_cols_e])):
            sheet = workbook[sheet_name]
            sheet.cell(row=1, column=1).fill = red_fill

            for i in range(c - 1):
                for j in highlight_cols[i]:
                    sheet.cell(row=j + 2, column=i + 2).fill = red_fill

            # 标红有问题的数据
            for row_idx, col_idx in problematic_cells:
                if col_idx <= c - 1:
                    sheet.cell(row=row_idx + 2, column=col_idx + 1).fill = problem_fill

        # 自动调整列宽
        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(output_path)
        workbook.close()

        return {
            'success': True,
            'output_file': output_filename,
            'problematic_data': problematic_data,
            'problematic_cells_count': len(problematic_cells),
            'attendance_issues': attendance_issues,
            'attendance_summary': {
                'late_count': sum(len(rows) for rows in highlight_cols_m),
                'no_lunch_count': sum(len(rows) for rows in highlight_cols_n),
                'early_leave_count': sum(len(rows) for rows in highlight_cols_e)
            },
            'employee_count': len(df_final),
            'total_working_hours': sum(Total_HEG),
            'total_overtime': sum(Total_OT)
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }

# HTML模板 - 包含修改后error表格上传功能
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>打卡数据处理系统</title>
    <link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>🕐</text></svg>">
    <style>
        body { font-family: Arial, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center; margin: 0; }
        .container { background: white; border-radius: 20px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); padding: 40px; max-width: 1000px; width: 95%; }
        .title { text-align: center; color: #333; margin-bottom: 30px; font-size: 2.5em; font-weight: bold; }
        .step { margin-bottom: 30px; padding: 20px; border: 2px solid #e1e5e9; border-radius: 10px; transition: all 0.3s ease; }
        .step.active { border-color: #667eea; background: #f8f9ff; }
        .step.completed { border-color: #28a745; background: #f8fff9; }
        .step-title { font-size: 1.3em; font-weight: bold; color: #333; margin-bottom: 15px; display: flex; align-items: center; }
        .step-number { background: #667eea; color: white; width: 30px; height: 30px; border-radius: 50%; display: flex; align-items: center; justify-content: center; margin-right: 10px; font-weight: bold; }
        .completed .step-number { background: #28a745; }
        .btn { background: #667eea; color: white; border: none; padding: 12px 24px; border-radius: 8px; cursor: pointer; font-size: 16px; transition: all 0.3s ease; margin-right: 10px; }
        .btn:hover { background: #5a6fd8; transform: translateY(-2px); }
        .btn:disabled { background: #ccc; cursor: not-allowed; transform: none; }
        .btn.success { background: #28a745; }
        .btn.warning { background: #ffc107; color: #333; }
        .file-label { display: inline-block; padding: 12px 24px; background: #667eea; color: white; border-radius: 8px; cursor: pointer; transition: background 0.3s ease; }
        .file-label:hover { background: #5a6fd8; }
        input[type="file"] { display: none; }
        .result { margin-top: 15px; padding: 15px; border-radius: 8px; display: none; }
        .result.success { background: #d4edda; border: 1px solid #c3e6cb; color: #155724; }
        .result.error { background: #f8d7da; border: 1px solid #f5c6cb; color: #721c24; }
        .loading { display: none; text-align: center; margin: 15px 0; }
        .spinner { border: 3px solid #f3f3f3; border-top: 3px solid #667eea; border-radius: 50%; width: 30px; height: 30px; animation: spin 1s linear infinite; margin: 0 auto 10px; }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
        .file-info { margin-top: 10px; padding: 10px; background: #e9ecef; border-radius: 5px; display: none; }
        .download-area { margin-top: 20px; text-align: center; }
        .error-details { max-height: 300px; overflow-y: auto; margin-top: 10px; padding: 15px; background: #fff3cd; border: 1px solid #ffeaa7; border-radius: 5px; font-size: 0.9em; }
        .error-item { margin: 5px 0; padding: 3px 0; border-bottom: 1px solid #ddd; }
        .summary-box { background: #e7f3ff; border: 1px solid #b8daff; border-radius: 5px; padding: 15px; margin: 10px 0; }
        .summary-item { display: inline-block; margin: 5px 15px 5px 0; font-weight: bold; }
        .icon { margin-right: 8px; }
        .collapsible { cursor: pointer; padding: 10px; background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 5px; margin: 5px 0; }
        .collapsible:hover { background: #e9ecef; }
        .content { display: none; padding: 15px; border: 1px solid #dee2e6; border-top: none; border-radius: 0 0 5px 5px; }
        .step-options { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 15px; }
    </style>
</head>
<body>
    <div class="container">
        <h1 class="title">🕐 打卡数据处理系统</h1>

        <div class="step active" id="step1">
            <div class="step-title">
                <span class="step-number">1</span>
                上传Timecard文件
            </div>
            <div>
                <label for="file" class="file-label">📁 选择Excel文件</label>
                <input type="file" id="file" accept=".xlsx,.xls">
                <div class="file-info" id="fileInfo"></div>
            </div>
            <button class="btn" id="uploadBtn" onclick="uploadFile()" disabled>上传文件</button>
            <div class="loading" id="uploadLoading"><div class="spinner"></div><div>正在上传...</div></div>
            <div class="result" id="uploadResult"></div>
        </div>

        <div class="step" id="step2">
            <div class="step-title">
                <span class="step-number">2</span>
                数据预处理与错误检测
            </div>
            <p>提取员工打卡数据，检测时间格式错误，生成高亮标记的错误检查表</p>
            <div class="step-options">
                <button class="btn" id="step1Btn" onclick="processStep1()" disabled>开始预处理</button>
                <button class="btn warning" id="uploadErrorBtn" onclick="uploadErrorFile()" disabled>上传修改后的错误表格</button>
            </div>
            <div class="loading" id="step1Loading"><div class="spinner"></div><div>正在处理...</div></div>
            <div class="result" id="step1Result"></div>
        </div>

        <div class="step" id="step3">
            <div class="step-title">
                <span class="step-number">3</span>
                工时计算与考勤分析
            </div>
            <p>计算工作时间、加班时间，检测迟到早退，生成完整的考勤报告</p>
            <button class="btn" id="step2Btn" onclick="processStep2()" disabled>生成报告</button>
            <div class="loading" id="step2Loading"><div class="spinner"></div><div>正在生成...</div></div>
            <div class="result" id="step2Result"></div>
        </div>

        <div class="step" id="step4">
            <div class="step-title">
                <span class="step-number">4</span>
                下载结果文件
            </div>
            <div class="download-area" id="downloadArea" style="display: none;">
                <button class="btn success" id="downloadErrorBtn" onclick="downloadFile()" style="display: none;">📥 下载错误检查表</button>
                <button class="btn success" id="downloadFinalBtn" onclick="downloadFinalFile()" style="display: none;">📊 下载最终报告</button>
            </div>
        </div>
    </div>

    <script>
        const API_BASE = '/api';
        let uploadedFilename = '', timeRange = '', errorFilename = '', finalFilename = '';

        document.getElementById('file').addEventListener('change', function(e) {
            const file = e.target.files[0];
            if (file) {
                document.getElementById('fileInfo').innerHTML = '<strong>已选择:</strong> ' + file.name + '<br><strong>大小:</strong> ' + (file.size / 1024 / 1024).toFixed(2) + ' MB';
                document.getElementById('fileInfo').style.display = 'block';
                document.getElementById('uploadBtn').disabled = false;
            }
        });

        async function uploadFile() {
            const file = document.getElementById('file').files[0];
            if (!file) { alert('请选择文件'); return; }

            const formData = new FormData();
            formData.append('file', file);

            showLoading('uploadLoading');
            try {
                const response = await fetch(API_BASE + '/upload', { method: 'POST', body: formData });
                const result = await response.json();
                hideLoading('uploadLoading');

                if (result.success) {
                    uploadedFilename = result.filename;
                    showResult('uploadResult', '文件上传成功！', 'success');
                    completeStep('step1');
                    activateStep('step2');
                    document.getElementById('step1Btn').disabled = false;
                } else {
                    showResult('uploadResult', '上传失败: ' + result.error, 'error');
                }
            } catch (error) {
                hideLoading('uploadLoading');
                showResult('uploadResult', '上传失败: ' + error.message, 'error');
            }
        }

        async function processStep1() {
            showLoading('step1Loading');
            try {
                const response = await fetch(API_BASE + '/process/step1', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ filename: uploadedFilename })
                });
                const result = await response.json();
                hideLoading('step1Loading');

                if (result.success) {
                    timeRange = result.time_range;
                    errorFilename = result.output_file;

                    let message = '<div class="summary-box">' +
                        '<div class="summary-item">👥 员工数量: ' + result.employee_count + '</div>' +
                        '<div class="summary-item">📅 时间范围: ' + timeRange + '</div>' +
                        '<div class="summary-item">🔍 错误数量: ' + result.error_count + '</div>' +
                        '<div class="summary-item">⚠️ 高亮单元格: ' + result.total_highlighted + '</div>' +
                        '</div>';

                    if (result.error_details && result.error_details.length > 0) {
                        message += '<div class="collapsible" onclick="toggleContent(\\'errorDetails1\\')">📋 错误详情 (点击展开)</div>' +
                            '<div id="errorDetails1" class="content">' +
                            '<div class="error-details">';
                        result.error_details.forEach(function(detail) {
                            message += '<div class="error-item">• ' + detail + '</div>';
                        });
                        message += '</div></div>';
                    }

                    message += '<br><strong>💡 提示：</strong>您可以下载错误检查表，修改后重新上传，或者直接继续下一步。';

                    showResult('step1Result', message, 'success');
                    document.getElementById('uploadErrorBtn').disabled = false;
                    document.getElementById('step2Btn').disabled = false;
                    document.getElementById('downloadErrorBtn').style.display = 'inline-block';
                    document.getElementById('downloadArea').style.display = 'block';
                } else {
                    showResult('step1Result', '处理失败: ' + result.error, 'error');
                }
            } catch (error) {
                hideLoading('step1Loading');
                showResult('step1Result', '处理失败: ' + error.message, 'error');
            }
        }

        async function uploadErrorFile() {
            // 创建文件输入元素
            const fileInput = document.createElement('input');
            fileInput.type = 'file';
            fileInput.accept = '.xlsx,.xls';
            fileInput.style.display = 'none';
            
            fileInput.onchange = async function(e) {
                const file = e.target.files[0];
                if (!file) return;

                const formData = new FormData();
                formData.append('file', file);
                formData.append('time_range', timeRange);

                showLoading('step1Loading');
                try {
                    const response = await fetch(API_BASE + '/upload/error', { method: 'POST', body: formData });
                    const result = await response.json();
                    hideLoading('step1Loading');

                    if (result.success) {
                        errorFilename = result.filename;
                        showResult('step1Result', '修改后的错误表格上传成功！现在可以继续下一步。', 'success');
                        document.getElementById('step2Btn').disabled = false;
                    } else {
                        showResult('step1Result', '上传失败: ' + result.error, 'error');
                    }
                } catch (error) {
                    hideLoading('step1Loading');
                    showResult('step1Result', '上传失败: ' + error.message, 'error');
                }
            };

            document.body.appendChild(fileInput);
            fileInput.click();
            document.body.removeChild(fileInput);
        }

        async function processStep2() {
            showLoading('step2Loading');
            try {
                const response = await fetch(API_BASE + '/process/step2', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ error_filename: errorFilename, time_range: timeRange })
                });
                const result = await response.json();
                hideLoading('step2Loading');

                if (result.success) {
                    finalFilename = result.output_file;

                    let message = '<div class="summary-box">' +
                        '<div class="summary-item">👥 处理员工: ' + result.employee_count + '</div>' +
                        '<div class="summary-item">⏰ 总工时: ' + (result.total_working_hours ? result.total_working_hours.toFixed(1) : 0) + 'h</div>' +
                        '<div class="summary-item">🔄 加班时间: ' + (result.total_overtime ? result.total_overtime.toFixed(1) : 0) + 'h</div>' +
                        '<div class="summary-item">🔥 问题数据: ' + result.problematic_cells_count + '</div>' +
                        '</div>';

                    if (result.attendance_summary) {
                        message += '<div class="summary-box">' +
                            '<div class="summary-item" style="color: #dc3545;">🐌 迟到: ' + result.attendance_summary.late_count + '次</div>' +
                            '<div class="summary-item" style="color: #fd7e14;">🍽️ 中午不打卡: ' + result.attendance_summary.no_lunch_count + '次</div>' +
                            '<div class="summary-item" style="color: #6f42c1;">🏃 早退: ' + result.attendance_summary.early_leave_count + '次</div>' +
                            '</div>';
                    }

                    message += '<br><strong>📊 Excel文件包含4个工作表：</strong><br>' +
                        '• 时间汇总：原始打卡时间 + 计算工时 + 统计数据<br>' +
                        '• 迟到：标红迟到记录<br>' +
                        '• 中午不打卡：标红中午不打卡记录<br>' +
                        '• 早退：标红早退记录<br><br>' +
                        '<strong>🎨 颜色说明：</strong><br>' +
                        '• 🟡 黄色：工时统计列<br>' +
                        '• 🔴 浅红色：迟到/早退/中午不打卡<br>' +
                        '• 🔥 深红色：有问题的时间数据 (需要人工确认)<br>' +
                        '• 🟢 绿色：假期列';

                    showResult('step2Result', message, 'success');
                    completeStep('step3');
                    completeStep('step4');
                    document.getElementById('downloadFinalBtn').style.display = 'inline-block';
                } else {
                    showResult('step2Result', '处理失败: ' + result.error, 'error');
                }
            } catch (error) {
                hideLoading('step2Loading');
                showResult('step2Result', '处理失败: ' + error.message, 'error');
            }
        }

        function downloadFile() {
            if (errorFilename) window.open(API_BASE + '/download/' + errorFilename, '_blank');
        }

        function downloadFinalFile() {
            if (finalFilename) window.open(API_BASE + '/download/' + finalFilename, '_blank');
        }

        function toggleContent(id) {
            const content = document.getElementById(id);
            content.style.display = content.style.display === 'block' ? 'none' : 'block';
        }

        function showLoading(id) { document.getElementById(id).style.display = 'block'; }
        function hideLoading(id) { document.getElementById(id).style.display = 'none'; }
        function showResult(id, message, type) {
            const element = document.getElementById(id);
            element.innerHTML = message;
            element.className = 'result ' + type;
            element.style.display = 'block';
        }
        function activateStep(stepId) { document.getElementById(stepId).classList.add('active'); }
        function completeStep(stepId) {
            const step = document.getElementById(stepId);
            step.classList.remove('active');
            step.classList.add('completed');
        }
    </script>
</body>
</html>
'''

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/favicon.ico')
def favicon():
    svg_icon = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100">
        <circle cx="50" cy="50" r="45" fill="#667eea" stroke="#5a6fd8" stroke-width="2"/>
        <text x="50" y="65" text-anchor="middle" font-size="40" fill="white">🕐</text>
    </svg>'''
    return svg_icon, 200, {'Content-Type': 'image/svg+xml'}

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400

    file = request.files['file']
    if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请选择Excel文件'}), 400

    filename = str(uuid.uuid4()) + '_' + file.filename
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)

    return jsonify({
        'success': True,
        'filename': filename,
        'original_name': file.filename
    })

@app.route('/api/upload/error', methods=['POST'])
def upload_error_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400

    file = request.files['file']
    if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请选择Excel文件'}), 400

    filename = str(uuid.uuid4()) + '_error_' + file.filename
    file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
    file.save(file_path)

    return jsonify({
        'success': True,
        'filename': filename,
        'original_name': file.filename
    })

@app.route('/api/process/step1', methods=['POST'])
def process_step1():
    data = request.json
    filename = data.get('filename')
    if not filename:
        return jsonify({'error': '缺少文件名'}), 400

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        return jsonify({'error': '文件不存在'}), 404

    result = process_timecard_step1(file_path)
    return jsonify(result)

@app.route('/api/process/step2', methods=['POST'])
def process_step2():
    data = request.json
    error_filename = data.get('error_filename')
    time_range = data.get('time_range')

    if not error_filename or not time_range:
        return jsonify({'error': '缺少必要参数'}), 400

    error_file_path = os.path.join(app.config['PROCESSED_FOLDER'], error_filename)
    if not os.path.exists(error_file_path):
        return jsonify({'error': '中间文件不存在'}), 404

    result = process_timecard_step2(error_file_path, time_range)
    return jsonify(result)

@app.route('/api/download/<filename>')
def download_file(filename):
    file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=filename)
    return jsonify({'error': '文件不存在'}), 404

@app.route('/api/status')
def status():
    return jsonify({
        'status': 'running',
        'upload_folder': app.config['UPLOAD_FOLDER'],
        'processed_folder': app.config['PROCESSED_FOLDER']
    })

if __name__ == '__main__':
    print("🚀 启动打卡数据处理系统...")
    print("📱 访问地址: http://localhost:8080")
    print("✨ 包含完整的错误检测、高亮标记和详细报告功能")
    print("🔄 新增：支持上传修改后的错误表格重新处理")
    app.run(host='0.0.0.0', port=8080, debug=True) 