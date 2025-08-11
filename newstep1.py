print('Hello! Welcome!')

import os
import re
import importlib.util

if importlib.util.find_spec('pandas') is None:
    print("Pandas not installed.  Please install the pandas library by typing 'pip install pandas'")
    quit()

if importlib.util.find_spec('numpy') is None:
    print("Numpy not installed.  Please install the numpy library by typing 'pip install numpy'")
    quit()

if importlib.util.find_spec('openpyxl') is None:
    print("openpyxl not installed.  Please install the openpyxl library by typing 'pip install openpyxl'")
    quit()

# 移除xlwings依赖检查，改用openpyxl来处理Excel格式

if importlib.util.find_spec('holidays') is None:
    print("holidays not installed.  Please install the holidays library by typing 'pip install holidays'")
    quit()

import pandas as pd
import numpy as np
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

directory = os.path.dirname(__file__)
print(directory)

# 查找Timecard文件
timecard_file = None
for file in os.listdir(directory):
    print(file)
    if 'Timecard' in file:
        import_path = os.path.join(directory, file)
        print(import_path)
        df = pd.read_excel(import_path)
        timecard_file = file
        break

if timecard_file is None:
    print("未找到Timecard文件!")
    quit()

# 获取时间范围
time_range = (df.iloc[1, 2]).replace("/", "").replace("~", "-").replace(" ", "")
file_name = 'table_with_error_cells' + '(' + time_range + ')' + '.xlsx'

output_path = os.path.join(directory, file_name)
print(output_path)

# 删除已存在的输出文件
if os.path.exists(output_path):
    os.remove(output_path)

total_rows = df.shape[0]
employee_amount = int((total_rows - 2) / 3)

date_row = df.iloc[2].to_list()
date_range = [x for x in date_row if str(x) != 'nan']
columns_name = list(map(int, date_range))
columns_name.insert(0, 'name')

df_new = pd.DataFrame(index=range(employee_amount), columns=columns_name)

# 创建新的员工和日常检查表
for i in range(employee_amount):
    df_new.iloc[i, 0] = df.iloc[(i + 1) * 3, 10]
    df_new.iloc[i, 1:] = df.iloc[(3 * (i + 1) + 1), 0:len(date_range)]

df_new['nan_count'] = df_new.isna().sum(axis=1)
df_new_sorted = df_new.sort_values(by='nan_count', ascending=True).reset_index(drop=True)
df_new = df_new_sorted.drop('nan_count', axis=1)


def get_minimum_distance(letter):
    position = []
    for i in range(len(letter)):
        if letter[i] == ':':
            position.append(i)
        else:
            continue

    if len(position) > 1:
        # 计算每个":"之间的距离
        j = len(position) - 1
        distance = []
        while j >= 1:
            d = position[j] - position[j - 1]
            j = j - 1
            distance.append(d)
        # 获取最小距离
        return min(distance)
    else:
        return None


# 通过计算':'之间的距离来获取有导入问题的值的位置
error_value_location = []
for i in range(employee_amount):
    for j in range(len(date_range)):
        string = str(df_new.iloc[i, j + 1])

        if string == 'nan':
            continue
        else:
            letter = [x for x in string]
            min_distance = get_minimum_distance(letter)
            if min_distance == 3:
                location = [i, j + 1]
                error_value_location.append(location)

highlight_index = []

for i in range(len(date_range)):
    time_rows = df_new[((df_new.iloc[:, i + 1].str.count(":") % 2) == 1).values].index.to_list()
    highlight_index.append(time_rows)

for i in range(len(error_value_location)):
    r = error_value_location[i][0]
    c = error_value_location[i][1]
    highlight_index[c - 1].append(r)

# 先保存Excel文件
df_new.to_excel(output_path, index=None, header=True)

# 使用openpyxl添加高亮显示
print("开始高亮显示差异...")
print("请稍候...")

workbook = load_workbook(output_path)
worksheet = workbook.active

# 定义高亮颜色
highlight_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# 高亮Excel中的行
for i in range(len(date_range)):
    for j in highlight_index[i]:
        # Excel中的行号从1开始，列号从1开始，我们需要+2因为有标题行
        cell = worksheet.cell(row=j + 2, column=i + 2)
        cell.fill = highlight_fill

workbook.save(output_path)
workbook.close()

print(f"处理完成! 输出文件: {output_path}")
print(f"发现 {len(error_value_location)} 个错误值位置")
print("高亮显示的单元格表示:")
print("- 红色背景: 时间记录有问题的单元格")