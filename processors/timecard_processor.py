import pandas as pd
import numpy as np
from datetime import datetime
import holidays
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.comments import Comment
import traceback
import os
from utils.time_utils import (
    get_minimum_distance,
    daily_working_time,
    parse_time_string,
    validate_time_format,
    normalize_time_list,
    detect_time_anomalies,
    calculate_working_hours_with_details
)


class TimecardProcessor:
    def __init__(self, upload_folder, processed_folder):
        self.upload_folder = upload_folder
        self.processed_folder = processed_folder

    def process_step1(self, file_path):
        """Step1处理逻辑 - 修复高亮显示问题"""
        try:
            df = pd.read_excel(file_path)
            print("📊 开始Step1处理...")
            print(f"📝 原始数据形状: {df.shape}")

            # 获取时间范围
            time_range = (df.iloc[1, 2]).replace("/", "").replace("~", "-").replace(" ", "")
            print(f"📅 时间范围: {time_range}")

            total_rows = df.shape[0]
            employee_amount = int((total_rows - 2) / 3)
            print(f"👥 员工数量: {employee_amount}")

            date_row = df.iloc[2].to_list()
            date_range = [x for x in date_row if str(x) != 'nan']
            columns_name = list(map(int, date_range))
            columns_name.insert(0, 'name')

            df_new = pd.DataFrame(index=range(employee_amount), columns=columns_name)

            # 创建新的员工和日常检查表
            for i in range(employee_amount):
                df_new.iloc[i, 0] = df.iloc[(i + 1) * 3, 10]
                df_new.iloc[i, 1:] = df.iloc[(3 * (i + 1) + 1), 0:len(date_range)]

            df_new['nan_count'] = df_new.isna().sum(axis=1)
            df_new_sorted = df_new.sort_values(by='nan_count', ascending=True).reset_index(drop=True)
            df_new = df_new_sorted.drop('nan_count', axis=1)

            # 增强的错误检测和高亮映射
            print("🔍 开始增强的错误检测...")

            # 用于存储每个单元格的异常信息
            cell_anomalies = {}  # key: (row, col), value: list of anomalies
            all_anomalies = []
            error_value_location = []

            # 定义异常类型颜色映射
            anomaly_colors = {
                'colon_distance': 'FFC7CE',  # 浅红色
                'odd_time_count': 'FF0000',  # 深红色
                'long_work_span': 'FFD700',  # 金色
                'time_sequence_error': 'FF8C00',  # 深橙色
                'invalid_time_format': 'FF6B6B',  # 橙红色
                'parse_error': '9932CC',  # 紫色
                'mixed_separators': '87CEEB',  # 天蓝色
                'default': 'FFC7CE'  # 默认浅红色
            }

            for i in range(employee_amount):
                employee_name = df_new.iloc[i, 0]
                for j in range(len(date_range)):
                    cell_value = str(df_new.iloc[i, j + 1])

                    if cell_value == 'nan':
                        continue

                    cell_key = (i, j + 1)
                    cell_anomalies[cell_key] = []

                    # 使用增强的异常检测
                    anomalies = detect_time_anomalies(cell_value, employee_name, j + 1)

                    if anomalies:
                        all_anomalies.extend(anomalies)
                        cell_anomalies[cell_key].extend(anomalies)

                        # 如果有严重错误，添加到错误位置列表
                        for anomaly in anomalies:
                            if anomaly['severity'] == 'error':
                                if [i, j + 1] not in error_value_location:
                                    error_value_location.append([i, j + 1])

                    # 原有的冒号距离检测（保持兼容性）
                    letter = [x for x in cell_value]
                    min_distance = get_minimum_distance(letter)
                    if min_distance == 3:
                        colon_anomaly = {
                            'type': 'colon_distance',
                            'message': f'冒号距离异常 - 员工: {employee_name}, 列: {j + 1}',
                            'severity': 'warning',
                            'color': 'FFC7CE',
                            'description': '时间格式问题，冒号前后数字位数异常'
                        }
                        if colon_anomaly not in cell_anomalies[cell_key]:
                            cell_anomalies[cell_key].append(colon_anomaly)
                        if [i, j + 1] not in error_value_location:
                            error_value_location.append([i, j + 1])

                    # 检测奇数时间记录
                    time_list = parse_time_string(cell_value)
                    if len(time_list) % 2 == 1:  # 奇数时间记录
                        odd_anomaly = {
                            'type': 'odd_time_count',
                            'message': f'奇数时间记录 - 员工: {employee_name}, 列: {j + 1}',
                            'severity': 'error',
                            'color': 'FF0000',
                            'description': '打卡次数为奇数，无法配对计算工时'
                        }
                        if odd_anomaly not in cell_anomalies[cell_key]:
                            cell_anomalies[cell_key].append(odd_anomaly)
                        if [i, j + 1] not in error_value_location:
                            error_value_location.append([i, j + 1])

            # 保存文件并添加高亮
            output_filename = f'table_with_error_cells({time_range}).xlsx'
            output_path = os.path.join(self.processed_folder, output_filename)
            df_new.to_excel(output_path, index=None, header=True)

            # 使用openpyxl添加高亮显示
            print("🎨 应用高亮显示...")
            workbook = load_workbook(output_path)
            worksheet = workbook.active

            # 应用高亮显示
            for cell_key, anomalies in cell_anomalies.items():
                if not anomalies:
                    continue

                row_idx, col_idx = cell_key

                # 确定要使用的颜色（优先级：error > warning）
                color = 'FFC7CE'  # 默认颜色
                comment_text = ""

                # 按严重程度选择颜色
                error_anomalies = [a for a in anomalies if a['severity'] == 'error']
                warning_anomalies = [a for a in anomalies if a['severity'] == 'warning']

                if error_anomalies:
                    # 优先显示错误级别的异常
                    primary_anomaly = error_anomalies[0]
                    color = anomaly_colors.get(primary_anomaly['type'], 'FF0000')
                    comment_text = primary_anomaly['description']
                elif warning_anomalies:
                    # 显示警告级别的异常
                    primary_anomaly = warning_anomalies[0]
                    color = anomaly_colors.get(primary_anomaly['type'], 'FFC7CE')
                    comment_text = primary_anomaly['description']

                # 添加所有异常到注释中
                if len(anomalies) > 1:
                    comment_text += f"\n\n共发现{len(anomalies)}个问题："
                    for i, anomaly in enumerate(anomalies, 1):
                        comment_text += f"\n{i}. {anomaly['description']}"

                # 应用高亮
                try:
                    # Excel行列索引从1开始，数据行从第2行开始（因为有标题行）
                    excel_row = row_idx + 2
                    excel_col = col_idx + 1

                    cell = worksheet.cell(row=excel_row, column=excel_col)
                    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.fill = fill

                    # 添加注释
                    if comment_text:
                        cell.comment = Comment(comment_text, "系统检测")

                    print(f"✅ 高亮单元格: 行{excel_row}, 列{excel_col}, 颜色#{color}")

                except Exception as e:
                    print(f"❌ 高亮失败: 行{row_idx}, 列{col_idx}, 错误: {e}")

            workbook.save(output_path)
            workbook.close()

            # 生成增强的错误报告
            error_details = []
            total_highlighted = len(cell_anomalies)

            # 按异常类型统计
            anomaly_stats = {}
            for anomaly in all_anomalies:
                anomaly_type = anomaly['type']
                if anomaly_type not in anomaly_stats:
                    anomaly_stats[anomaly_type] = 0
                anomaly_stats[anomaly_type] += 1

            type_names = {
                'colon_distance': '冒号距离异常',
                'odd_time_count': '奇数时间记录',
                'long_work_span': '工作时间跨度异常',
                'time_sequence_error': '时间顺序错误',
                'invalid_time_format': '无效时间格式',
                'parse_error': '解析错误',
                'mixed_separators': '混合分隔符'
            }

            for anomaly_type, count in anomaly_stats.items():
                type_name = type_names.get(anomaly_type, anomaly_type)
                error_details.append(f"发现 {count} 个{type_name}")

            print(f"✅ Step1处理完成")
            print(f"📊 统计信息:")
            print(f"   - 员工数量: {employee_amount}")
            print(f"   - 错误位置: {len(error_value_location)}")
            print(f"   - 高亮单元格: {total_highlighted}")
            print(f"   - 异常类型: {list(anomaly_stats.keys())}")

            return {
                'success': True,
                'time_range': time_range,
                'output_file': output_filename,
                'employee_count': employee_amount,
                'error_count': len(error_value_location),
                'total_highlighted': total_highlighted,
                'error_details': error_details,
                'anomaly_stats': anomaly_stats
            }

        except Exception as e:
            print(f"❌ Step1处理失败: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'traceback': traceback.format_exc()
            }

    def process_step2(self, error_file_path, time_range):
        """Step2处理逻辑 - 修复行列对齐问题"""
        try:
            print("📊 开始Step2处理...")
            df = pd.read_excel(error_file_path)
            df_new = df.copy()

            # 保存原始打卡时间数据（用于显示）
            df_original_times = df.copy()
            print(f"📝 原始数据形状: {df_original_times.shape}")
            print(f"📝 列名: {list(df_original_times.columns)}")

            # 转换为字符串进行处理
            df = df.astype(str)

            # 记录有问题的数据和单元格位置
            problematic_data = []
            problematic_cells_with_details = {}  # key: (row, col), value: anomaly_info
            processing_stats = {
                'total_cells': 0,
                'valid_cells': 0,
                'invalid_cells': 0,
                'zero_hour_cells': 0
            }

            print("🔄 开始时间数据处理和工时计算...")

            # 获取原始数据的基本信息
            num_employees = len(df)
            num_date_cols = len(df.columns) - 1  # 减去name列
            print(f"👥 员工数量: {num_employees}")
            print(f"📅 日期列数: {num_date_cols}")

            # 处理时间数据并计算工时
            for i in range(num_employees):
                employee_name = df.iloc[i, 0]
                print(f"🔄 处理员工: {employee_name} (行 {i})")

                for j in range(num_date_cols):
                    processing_stats['total_cells'] += 1

                    if (df.iloc[i, j + 1] == 'nan'):
                        continue

                    raw_time_str = str(df.iloc[i, j + 1])

                    # 检测异常
                    anomalies = detect_time_anomalies(raw_time_str, employee_name, j + 1)

                    # 使用增强的时间解析
                    time_list = parse_time_string(raw_time_str)
                    time_list_normalized = normalize_time_list(time_list)

                    # 使用增强的工时计算
                    work_result = calculate_working_hours_with_details(time_list_normalized)

                    # 确定异常类型和描述
                    anomaly_type = None
                    anomaly_description = ""
                    anomaly_color = 'FF0000'  # 默认深红色

                    if anomalies:
                        # 优先使用检测到的异常
                        primary_anomaly = anomalies[0]
                        anomaly_type = primary_anomaly['type']
                        anomaly_description = primary_anomaly['description']
                        anomaly_color = primary_anomaly.get('color', 'FF0000')
                    elif not work_result['is_valid']:
                        anomaly_type = 'calculation_error'
                        anomaly_description = f"计算错误: {work_result['error']}"
                        anomaly_color = 'FF4500'  # 橙红色
                    elif work_result['total_hours'] == 0 and len(time_list_normalized) > 0:
                        anomaly_type = 'zero_hours'
                        anomaly_description = '工时为零'
                        anomaly_color = 'FFB6C1'  # 浅粉色
                    elif work_result['total_hours'] > 12:
                        anomaly_type = 'long_work_span'
                        anomaly_description = f'工作时间异常长 ({work_result["total_hours"]}h)'
                        anomaly_color = 'FFD700'  # 金色

                    if work_result['is_valid']:
                        df_new.iloc[i, j + 1] = work_result['total_hours']
                        processing_stats['valid_cells'] += 1

                        print(f"   ✅ 列{j + 1}: {raw_time_str} -> {work_result['total_hours']}h")

                        # 如果有异常，记录到问题数据中
                        if anomaly_type:
                            problematic_data.append(
                                f"{anomaly_description} - 员工: {employee_name}, 列: {j + 1}, "
                                f"工时: {work_result['total_hours']}h"
                            )
                            problematic_cells_with_details[(i, j + 1)] = {
                                'type': anomaly_type,
                                'description': anomaly_description,
                                'color': anomaly_color,
                                'raw_value': raw_time_str,
                                'employee': employee_name,
                                'column': j + 1,
                                'work_hours': work_result['total_hours']
                            }
                    else:
                        problematic_data.append(
                            f"{work_result['error']} - 员工: {employee_name}, 列: {j + 1}"
                        )
                        problematic_cells_with_details[(i, j + 1)] = {
                            'type': 'calculation_error',
                            'description': work_result['error'],
                            'color': 'FF4500',
                            'raw_value': raw_time_str,
                            'employee': employee_name,
                            'column': j + 1
                        }
                        df_new.iloc[i, j + 1] = 0
                        processing_stats['invalid_cells'] += 1

                        print(f"   ❌ 列{j + 1}: {raw_time_str} -> 计算失败")

                    if df_new.iloc[i, j + 1] == 0:
                        processing_stats['zero_hour_cells'] += 1

            print(f"📊 处理统计:")
            print(f"   - 总单元格: {processing_stats['total_cells']}")
            print(f"   - 有效单元格: {processing_stats['valid_cells']}")
            print(f"   - 无效单元格: {processing_stats['invalid_cells']}")
            print(f"   - 零工时单元格: {processing_stats['zero_hour_cells']}")

            # 计算工时统计
            print("📊 计算工时统计...")

            # 确保df_new中的数据是数值型
            for col in df_new.columns[1:]:  # 跳过name列
                df_new[col] = pd.to_numeric(df_new[col], errors='coerce').fillna(0)

            # 第一周工时计算（假设前7列是第一周）
            if num_date_cols >= 7:
                total1 = df_new.iloc[:, 1:8].sum(axis=1).to_list()
            else:
                total1 = df_new.iloc[:, 1:num_date_cols + 1].sum(axis=1).to_list()

            HEG1 = [min(40, max(0, t)) if t > 0 else 0 for t in total1]
            OT1 = [max(0, t - 40) if t > 40 else 0 for t in total1]

            # 第二周工时计算
            if num_date_cols >= 16:
                total2 = df_new.iloc[:, 8:17].sum(axis=1).to_list()
            elif num_date_cols > 7:
                total2 = df_new.iloc[:, 8:num_date_cols + 1].sum(axis=1).to_list()
            else:
                total2 = [0] * num_employees

            HEG2 = [min(40, max(0, t)) if t > 0 else 0 for t in total2]
            OT2 = [max(0, t - 40) if t > 40 else 0 for t in total2]

            # 总计
            Total_HEG = [HEG1[i] + HEG2[i] for i in range(num_employees)]
            Total_OT = [OT1[i] + OT2[i] for i in range(num_employees)]

            print("📋 构建最终显示数据框...")

            # 重新构建最终数据框，使用更简单的方法避免类型错误
            # 1. 从原始时间数据开始
            df_final = df_original_times.astype(str).replace('nan', '')

            print(f"📝 最终数据框初始形状: {df_final.shape}")
            print(f"📝 最终数据框列名: {list(df_final.columns)}")
            print(f"📝 列名类型: {[type(col) for col in df_final.columns]}")

            # 2. 简化方法：直接添加所有工时列到最后
            original_date_cols = len(df_final.columns) - 1  # 减去name列

            for i in range(1, min(len(df_final.columns), len(df_new.columns))):
                try:
                    date_col = df_final.columns[i]
                    hour_col = f"{date_col}_小时"
                    work_hours = df_new.iloc[:, i].values
                    df_final[hour_col] = work_hours
                    print(f"   ✅ 添加 {hour_col} 列")
                except Exception as e:
                    print(f"   ❌ 添加工时列失败: 列{i}, 错误: {e}")
                    continue

            # 3. 添加统计列到最后
            df_final["HEG1"] = HEG1
            df_final["OT1"] = OT1
            df_final["HEG2"] = HEG2
            df_final["OT2"] = OT2
            df_final["Total_HEG"] = Total_HEG
            df_final["Total_OT"] = Total_OT

            print(f"📝 最终数据框完成形状: {df_final.shape}")
            print(f"📝 最终数据框列名: {list(df_final.columns)}")

            # 4. 替换0为空字符串（仅在工时和统计列中）
            # 找出所有工时列和统计列的索引
            hour_cols = [col for col in df_final.columns if isinstance(col, str) and '_小时' in str(col)]
            stat_cols = ["HEG1", "OT1", "HEG2", "OT2", "Total_HEG", "Total_OT"]

            for col in hour_cols + stat_cols:
                if col in df_final.columns:
                    df_final[col] = df_final[col].replace(0, '')

            print("🕐 检测考勤问题...")
            # 识别需要检查迟到早退的员工
            name_list = []
            for i in range(num_employees):
                if HEG1[i] > 30 or HEG2[i] > 30:
                    name_list.append(df_final.iloc[i, 0])

            print(f"👥 需要检查考勤的员工: {len(name_list)} 人")

            # 处理原始时间数据用于考勤检测
            df_original_for_display = df_original_times.astype(str).replace('nan', '')

            # 检测迟到早退
            attendance_result = self._detect_attendance_issues_enhanced(df_original_for_display, name_list, df_final)

            # 处理假期
            print("🏖️ 处理假期信息...")
            holiday_result = self._process_holidays(time_range, df_final)

            # 创建Excel文件
            output_filename = f'work_attendance({time_range}).xlsx'
            output_path = os.path.join(self.processed_folder, output_filename)

            print("📋 生成Excel报告...")

            # 修正problematic_cells的列索引，只针对原始时间列
            corrected_problematic_cells = {}
            for (row_idx, col_idx), details in problematic_cells_with_details.items():
                # col_idx是基于df_new的列索引，需要确保只处理原始时间列
                if col_idx <= original_date_cols:  # 确保是原始时间列
                    corrected_problematic_cells[(row_idx, col_idx)] = details

            self._create_excel_report_enhanced(df_final, df_original_for_display, attendance_result,
                                               corrected_problematic_cells, original_date_cols,
                                               output_path, holiday_result, processing_stats)

            print(f"✅ Step2处理完成")
            print(f"📊 最终统计:")
            print(f"   - 总工时: {sum(Total_HEG):.1f}h")
            print(f"   - 加班时间: {sum(Total_OT):.1f}h")
            print(f"   - 问题单元格: {len(corrected_problematic_cells)}")

            return {
                'success': True,
                'output_file': output_filename,
                'problematic_data': problematic_data,
                'problematic_cells_count': len(corrected_problematic_cells),
                'attendance_issues': attendance_result['attendance_issues'],
                'attendance_summary': attendance_result['attendance_summary'],
                'employee_count': len(df_final),
                'total_working_hours': sum(Total_HEG),
                'total_overtime': sum(Total_OT),
                'processing_stats': processing_stats
            }

        except Exception as e:
            print(f"❌ Step2处理失败: {str(e)}")
            print(f"❌ 错误详情: {traceback.format_exc()}")
            return {
                'success': False,
                'error': str(e),
                'traceback': traceback.format_exc()
            }

    def _detect_attendance_issues_enhanced(self, df_original_for_display, name_list, df_final):
        """增强的考勤问题检测"""
        r = df_original_for_display.shape[0]
        c = df_original_for_display.shape[1]

        highlight_cols_m = []  # 迟到
        highlight_cols_n = []  # 中午不打卡
        highlight_cols_e = []  # 早退

        attendance_issues = []

        for i in range(c - 1):
            highlight_rows_m = []
            highlight_rows_n = []
            highlight_rows_e = []

            for j in range(r):
                if df_original_for_display.iloc[j, 0] in name_list:
                    if df_original_for_display.iloc[j, i + 1] == '':
                        continue

                    raw_time_str = str(df_original_for_display.iloc[j, i + 1])

                    # 使用增强的时间解析
                    time_list = parse_time_string(raw_time_str)
                    valid_times = normalize_time_list(time_list)

                    if len(valid_times) == 0:
                        continue

                    try:
                        check_in_time = valid_times[0]
                        check_out_time = valid_times[-1]
                        morning_reference_time = datetime.strptime('10:00', '%H:%M')
                        evening_reference_time = datetime.strptime('17:00', '%H:%M')
                        check_times = len(valid_times)

                        employee_name = df_original_for_display.iloc[j, 0]
                        date_col = df_original_for_display.columns[i + 1]

                        # 检测迟到
                        if check_in_time.time() > morning_reference_time.time():
                            highlight_rows_m.append(j)
                            attendance_issues.append(
                                f"迟到 - {employee_name}, {date_col}, 上班时间: {check_in_time.strftime('%H:%M')}")

                        # 检测中午不打卡
                        if check_times == 2:
                            highlight_rows_n.append(j)
                            attendance_issues.append(
                                f"中午不打卡 - {employee_name}, {date_col}, 打卡次数: {check_times}")

                        # 检测早退
                        if check_out_time.time() < evening_reference_time.time():
                            highlight_rows_e.append(j)
                            attendance_issues.append(
                                f"早退 - {employee_name}, {date_col}, 下班时间: {check_out_time.strftime('%H:%M')}")

                    except Exception as e:
                        print(f"⚠️ 考勤检测异常: 员工 {df_original_for_display.iloc[j, 0]}, 列 {i + 1}, 错误: {e}")
                        continue

            highlight_cols_m.append(highlight_rows_m)
            highlight_cols_n.append(highlight_rows_n)
            highlight_cols_e.append(highlight_rows_e)

        return {
            'highlight_cols_m': highlight_cols_m,
            'highlight_cols_n': highlight_cols_n,
            'highlight_cols_e': highlight_cols_e,
            'attendance_issues': attendance_issues,
            'attendance_summary': {
                'late_count': sum(len(rows) for rows in highlight_cols_m),
                'no_lunch_count': sum(len(rows) for rows in highlight_cols_n),
                'early_leave_count': sum(len(rows) for rows in highlight_cols_e)
            }
        }

    def _process_holidays(self, time_range, df_final):
        """处理假期信息 - 修复类型错误"""
        try:
            US_holidays = pd.DataFrame.from_dict(holidays.US(years=2022).items())
            US_holidays.columns = ["date", "holiday_name"]
            my_vacation = ["New Year's Day", "Independence Day", "Labor Day", "Thanksgiving", "Christmas Day"]

            temp_list = []
            for i in US_holidays['holiday_name']:
                if i in my_vacation:
                    x = US_holidays[US_holidays['holiday_name'] == i].index[0]
                    temp_list.append(x)

            my_vacation_date = list(US_holidays.iloc[temp_list, 0])

            # 解析时间范围
            time_str = time_range.split("-")
            if len(time_str[0]) != len(time_str[1]):
                year1 = time_str[0][:4]
                start_day = time_str[0][-4:]
                end_day = time_str[1][-4:]
                start_date = datetime.strptime(year1 + start_day, '%Y%m%d').date()
                end_date = datetime.strptime(year1 + end_day, '%Y%m%d').date()
            else:
                year1 = time_str[0][:4]
                year2 = time_str[1][:4]
                start_day = time_str[0][-4:]
                end_day = time_str[1][-4:]
                start_date = datetime.strptime(year1 + start_day, '%Y%m%d').date()
                end_date = datetime.strptime(year2 + end_day, '%Y%m%d').date()

            # 检查假期
            holiday_column = None
            for time in my_vacation_date:
                if start_date < time < end_date:
                    index = US_holidays[US_holidays['date'] == time].index[0]
                    holiday = US_holidays.holiday_name[index]
                    # 安全的列名检查和重命名
                    day_num = time.day
                    day_str = str(day_num)

                    renamed = False
                    # 检查所有可能的列名格式
                    for col in df_final.columns:
                        if str(col) == day_str or col == day_num:
                            df_final = df_final.rename(columns={col: holiday})
                            holiday_column = df_final.columns.get_loc(holiday)
                            renamed = True
                            break

                    if renamed:
                        print(f"📅 重命名假期列: {day_num} -> {holiday}")

            return {'holiday_column': holiday_column, 'df_final': df_final}
        except Exception as e:
            print(f"⚠️ 假期处理失败: {e}")
            return {'holiday_column': None, 'df_final': df_final}

    def _create_excel_report_enhanced(self, df_final, df_original_for_display, attendance_result,
                                      problematic_cells_with_details, original_date_cols,
                                      output_path, holiday_result, processing_stats):
        """创建增强的Excel报告 - 修复行列对齐"""
        workbook = Workbook()
        workbook.remove(workbook.active)

        # 创建工作表
        sheet_names = ["时间汇总", "迟到", "中午不打卡", "早退", "处理日志"]
        sheets_data = [df_final, df_original_for_display, df_original_for_display, df_original_for_display, None]

        for i, (sheet_name, data) in enumerate(zip(sheet_names, sheets_data)):
            ws = workbook.create_sheet(title=sheet_name)

            if data is not None:
                print(f"📝 写入工作表 '{sheet_name}', 形状: {data.shape}")
                for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
            else:
                # 处理日志工作表
                self._create_log_sheet(ws, processing_stats, attendance_result, problematic_cells_with_details)

        # 应用样式
        print("🎨 应用Excel样式和高亮...")
        self._apply_enhanced_styles_fixed(workbook, df_final, attendance_result,
                                          problematic_cells_with_details, original_date_cols,
                                          holiday_result)

        workbook.save(output_path)
        workbook.close()
        print(f"📁 Excel文件已保存: {output_path}")

    def _apply_enhanced_styles_fixed(self, workbook, df_final, attendance_result,
                                     problematic_cells_with_details, original_date_cols,
                                     holiday_result):
        """修复的高亮样式应用方法"""
        # 定义颜色
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        log_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        print("🎨 处理时间汇总工作表...")
        sheet1 = workbook["时间汇总"]

        # 高亮统计列（最后6列）
        total_cols = len(df_final.columns)
        metrics_start = total_cols - 6
        for i in range(metrics_start, total_cols):
            cell = sheet1.cell(row=1, column=i + 1)
            cell.fill = yellow_fill
            print(f"   ✅ 统计列高亮: 列{i + 1}")

        # 高亮员工姓名列
        sheet1.cell(row=1, column=1).fill = red_fill

        # 高亮假期列
        if holiday_result['holiday_column'] is not None:
            sheet1.cell(row=1, column=holiday_result['holiday_column'] + 1).fill = green_fill

        # 高亮问题数据单元格（只在原始时间列中）
        problem_count = 0
        for cell_key, anomaly_info in problematic_cells_with_details.items():
            row_idx, col_idx = cell_key

            # 确保只在原始时间列中高亮
            if col_idx <= original_date_cols:
                try:
                    excel_row = row_idx + 2  # +2 因为Excel从1开始且有标题行
                    excel_col = col_idx + 1  # +1 因为Excel从1开始

                    cell = sheet1.cell(row=excel_row, column=excel_col)

                    # 根据异常类型选择颜色
                    color = anomaly_info.get('color', 'FF0000')
                    problem_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.fill = problem_fill

                    # 添加注释
                    comment_text = f"异常类型: {anomaly_info['type']}\n"
                    comment_text += f"描述: {anomaly_info['description']}\n"
                    comment_text += f"员工: {anomaly_info['employee']}\n"
                    comment_text += f"原始值: {anomaly_info['raw_value']}"

                    if 'work_hours' in anomaly_info:
                        comment_text += f"\n计算工时: {anomaly_info['work_hours']}h"

                    cell.comment = Comment(comment_text, "系统检测")

                    problem_count += 1
                    print(f"   ✅ 高亮问题单元格: 行{excel_row}, 列{excel_col}, 类型: {anomaly_info['type']}")

                except Exception as e:
                    print(f"   ❌ 高亮失败: {cell_key}, 错误: {e}")

        print(f"📊 时间汇总工作表: 共高亮 {problem_count} 个问题单元格")

        # 处理考勤工作表
        attendance_sheets = [
            ("迟到", attendance_result['highlight_cols_m']),
            ("中午不打卡", attendance_result['highlight_cols_n']),
            ("早退", attendance_result['highlight_cols_e'])
        ]

        for sheet_name, highlight_cols in attendance_sheets:
            print(f"🎨 处理{sheet_name}工作表...")
            sheet = workbook[sheet_name]

            # 高亮标题
            sheet.cell(row=1, column=1).fill = red_fill

            # 高亮考勤问题单元格
            attendance_count = 0
            for col_idx, rows in enumerate(highlight_cols):
                for row_idx in rows:
                    try:
                        excel_row = row_idx + 2  # +2 因为Excel从1开始且有标题行
                        excel_col = col_idx + 2  # +2 因为第一列是姓名列，Excel从1开始

                        cell = sheet.cell(row=excel_row, column=excel_col)
                        cell.fill = red_fill

                        # 添加考勤问题注释
                        cell.comment = Comment(f"{sheet_name} - 需要关注", "系统检测")
                        attendance_count += 1

                    except Exception as e:
                        print(f"   ❌ 考勤高亮失败: 行{row_idx}, 列{col_idx}, 错误: {e}")

            print(f"📊 {sheet_name}工作表: 共高亮 {attendance_count} 个考勤问题")

        # 处理日志工作表样式
        print("🎨 处理日志工作表...")
        log_sheet = workbook["处理日志"]

        # 高亮标题行
        for row in range(1, log_sheet.max_row + 1):
            cell = log_sheet.cell(row=row, column=1)
            if cell.value and isinstance(cell.value, str) and ('统计' in cell.value or '说明' in cell.value):
                cell.fill = log_header_fill
                log_sheet.cell(row=row, column=2).fill = log_header_fill

        # 自动调整列宽
        print("📐 自动调整列宽...")
        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                # 设置合适的列宽
                adjusted_width = min((max_length + 2) * 1.2, 50)
                sheet.column_dimensions[column_letter].width = max(adjusted_width, 10)

        print("✅ 所有样式应用完成")

    def _create_log_sheet(self, ws, processing_stats, attendance_result, problematic_cells_with_details):
        """创建处理日志工作表"""
        log_data = [
            ["处理统计", ""],
            ["总单元格数", processing_stats['total_cells']],
            ["有效单元格数", processing_stats['valid_cells']],
            ["无效单元格数", processing_stats['invalid_cells']],
            ["零工时单元格数", processing_stats['zero_hour_cells']],
            ["", ""],
            ["考勤统计", ""],
            ["迟到次数", attendance_result['attendance_summary']['late_count']],
            ["中午不打卡次数", attendance_result['attendance_summary']['no_lunch_count']],
            ["早退次数", attendance_result['attendance_summary']['early_leave_count']],
            ["", ""],
            ["问题数据统计", ""],
            ["问题单元格总数", len(problematic_cells_with_details)],
            ["", ""],
            ["异常类型颜色说明", ""],
            ["冒号距离异常", "浅红色 #FFC7CE"],
            ["奇数时间记录", "深红色 #FF0000"],
            ["时间顺序错误", "深橙色 #FF8C00"],
            ["时间格式无效", "橙红色 #FF6B6B"],
            ["解析错误", "紫色 #9932CC"],
            ["混合分隔符", "天蓝色 #87CEEB"],
            ["工作时间跨度异常", "金色 #FFD700"],
            ["计算错误", "橙红色 #FF4500"],
            ["零工时", "浅粉色 #FFB6C1"]
        ]

        # 统计异常类型
        anomaly_stats = {}
        for anomaly_info in problematic_cells_with_details.values():
            anomaly_type = anomaly_info['type']
            if anomaly_type not in anomaly_stats:
                anomaly_stats[anomaly_type] = 0
            anomaly_stats[anomaly_type] += 1

        # 添加异常统计
        if anomaly_stats:
            log_data.extend([["", ""], ["各类型异常统计", ""]])
            for anomaly_type, count in anomaly_stats.items():
                log_data.append([anomaly_type, count])

        # 写入数据
        for r_idx, row in enumerate(log_data, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)